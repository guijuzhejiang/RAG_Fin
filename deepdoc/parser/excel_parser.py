#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#

import xlrd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import shutil
import traceback
from io import BytesIO
from spire.xls import *
from spire.xls.common import *
from rag.nlp import find_codec
from .utils import extract_table_from_html, AsposeConverter


class RAGFlowExcelParser:
    def html(self, fnm, chat_mdl, chunk_rows=8, overlap=2, type="xlsx"):
# merge_inferred_with_extracted_layout
        tb_chunks = []
        sheet_htmls = AsposeConverter.SaveToHtml(fnm, chat_mdl)
        for sheetname, buffer_path in sheet_htmls:
            try:
                tables = extract_table_from_html(buffer_path)
                soup = BeautifulSoup(features="html.parser")
                # 找到所有的行
                rows = [row for table in tables for row in table.find_all("tr")]
                # 根据chunk_rows分块处理数据行
                for rows_i in range(0, len(rows), chunk_rows - overlap):
                    tb = soup.new_tag("table")
                    # add the caption to the table
                    if sheetname and not sheetname.startswith('Evaluation Warning'):
                        caption = soup.new_tag("caption")
                        caption.string = sheetname
                        tb.append(caption)

                    # 确保块的起始和结束索引符合预期
                    start_index = rows_i
                    end_index = min(rows_i + chunk_rows, len(rows))  # 保证每块有固定的 chunk_rows 行，且不超出长度

                    # dispatch row
                    for r in list(rows[start_index:end_index]):
                        tb.append(r)

                    tb_chunks.append(str(tb).replace("\n", ""))

            except Exception:
                print(f"Error: {traceback.format_exc()}")
            finally:
                if os.path.exists(buffer_path):
                    os.remove(buffer_path)
                    dir_path = buffer_path.replace(".html", "_files")
                    if os.path.exists(dir_path):
                        shutil.rmtree(dir_path)

        return tb_chunks

    def __call__(self, fnm, type="xlsx"):
        res = []

        if type == "xlsx":
            if isinstance(fnm, str):
                wb = load_workbook(fnm)
            else:
                wb = load_workbook(BytesIO(fnm))
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                rows = list(ws.rows)
                if not rows:
                    continue
                ti = list(rows[0])
                for r in list(rows[1:]):
                    l_list = []
                    for i, c in enumerate(r):
                        if not c.value:
                            continue
                        t = str(ti[i].value) if i < len(ti) else ""
                        t += ("：" if t else "") + str(c.value)
                        l_list.append(t)
                    l_string = "; ".join(l_list)
                    if sheetname.lower().find("sheet") < 0:
                        l_string += " ——" + sheetname
                    res.append(l_string)
        else:
            # xls
            if isinstance(fnm, str):
                wb = xlrd.open_workbook(fnm)
            else:
                wb = xlrd.open_workbook(file_contents=BytesIO(fnm).read())

            for sheetname in wb.sheet_names():
                ws = wb.sheet_by_name(sheetname)
                rows = ws.get_rows()

                # 将所有行转换为列表
                rows = list(rows)

                # 如果工作表为空，跳过
                if not rows:
                    continue

                # 第一行为表头
                ti = [cell.value for cell in rows[0]]

                # 遍历工作表中的数据行
                for r in rows[1:]:
                    l_list = []
                    for i, c in enumerate(r):
                        if not c.value:
                            continue
                        t = str(ti[i]) if i < len(ti) else ""
                        t += ("：" if t else "") + str(c.value)
                        l_list.append(t)
                    l_string = "; ".join(l_list)
                    if sheetname.lower().find("sheet") < 0:
                        l_string += " ——" + sheetname
                    res.append(l_string)
        return res

    @staticmethod
    def row_number(fnm, binary):
        type_name = fnm.split(".")[-1].lower()

        if type_name.find("xls") >= 0:
            total = 0

            if type_name == "xlsx":
                wb = load_workbook(BytesIO(binary))
                for sheetname in wb.sheetnames:
                    ws = wb[sheetname]
                    total += len(list(ws.rows))
                    return total

            else:
                wb = xlrd.open_workbook(file_contents=BytesIO(binary).read())
                for sheetname in wb.sheet_names():
                    ws = wb.sheet_by_name(sheetname)
                    total += ws.nrows
                    return total

        if type_name in ["csv", "txt"]:
            encoding = find_codec(binary)
            txt = binary.decode(encoding, errors="ignore")
            return len(txt.split("\n"))


if __name__ == "__main__":
    psr = RAGFlowExcelParser()
    psr(sys.argv[1])
